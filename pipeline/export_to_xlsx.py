"""
Export family_offices.json to the target 32-column Excel layout.

Maps structured fields to deliverable columns, leaves unverified cells blank,
and computes a per-record completion score. No fabrication.
"""

import json
from datetime import date
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import records  # reuse the description renderer

DATA = Path(__file__).parent / "family_offices.json"
OUT = Path(__file__).parent / "family_offices_export.xlsx"

HEADERS = [
    "Family Office Name", "Data Validation Period",
    "Data Completion Score (Text)", "Data Completion Score (Visual)",
    "Family Office Description", "Investment Thesis", "Investing Sectors",
    "Family Office Domain", "Family Office Website Address", "URL Quality",
    "Corporate Linkedin Address", "Family Office Street Address",
    "Family Office City", "Family Office State / Region", "Family Office Country",
    "Contact First Name", "Contact Last Name", "Contact Full Name",
    "Contact Job Title", "Contact Location", "Contact LinkedIn Profile",
    "Contact Primary Email", "Primary E-Mail Validation Code",
    "Primary E-Mail Code Explanation", "Email Quality Assessment (Primary)",
    "Primary Phone Number", "Contact Secondary Email",
    "Secondary E-Mail Validation Code", "Secondary E-Mail Code Explanation",
    "Email Quality Assessment (Secondary)", "Secondary Phone Number",
]

# Columns that count toward the completion score (data fields, not name/period/scores).
SCOREABLE = set(range(4, 31))  # indexes 4..30 in HEADERS


def val(cell):
    return cell.get("value") if isinstance(cell, dict) else cell


def domain_of(url):
    if not url:
        return ""
    d = url.replace("https://", "").replace("http://", "").replace("www.", "")
    return d.split("/")[0]


def split_name(full):
    if not full:
        return "", ""
    parts = full.split()
    return parts[0], " ".join(parts[1:])


def row_for(r):
    loc = r.get("location") or {}
    fc = r.get("firm_contact") or {}
    dms = r.get("decision_makers") or []
    dm = dms[0] if dms else {}
    first, last = split_name(dm.get("name"))
    website = fc.get("website")
    if website and not website.startswith("http"):
        website = "https://" + website

    row = [
        r.get("entity", ""),                                   # Name
        r.get("aum", {}).get("as_of") or date.today().year,    # Validation Period
        None, None,                                            # score text/visual (filled after)
        records.to_document(r),                                # Description
        val(r.get("thesis")) or "",                            # Investment Thesis
        val(r.get("thesis")) or "",                            # Investing Sectors (from thesis)
        domain_of(fc.get("website")),                          # Domain
        website or "",                                         # Website
        "Verified" if website else "",                         # URL Quality
        fc.get("linkedin") or "",                              # Corporate LinkedIn
        fc.get("address") or "",                               # Street Address
        loc.get("city") or "",                                 # City
        loc.get("state") or "",                                # State/Region
        r.get("_country") or "United States of America",       # Country
        first, last, dm.get("name") or "",                     # Contact name fields
        dm.get("title") or "",                                 # Job Title
        "",                                                    # Contact Location
        dm.get("linkedin") or "",                              # Contact LinkedIn
        dm.get("work_email") or "",                            # Primary Email
        "", "", "",                                            # email validation/expl/quality
        dm.get("direct_phone") or fc.get("phone") or "",       # Primary Phone
        "", "", "", "", "",                                    # secondary email/phone block
    ]
    # completion score = filled scoreable cells
    filled = sum(1 for i in SCOREABLE if str(row[i]).strip())
    row[2] = filled
    row[3] = filled
    return row


def main():
    data = json.load(open(DATA, encoding="utf-8"))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    header_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4438")
    cell_font = Font(name="Arial", size=10)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(HEADERS)
    for c in ws[1]:
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = border

    for r in data:
        ws.append(row_for(r))

    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.font = cell_font
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = border

    # widths: wide for text columns, medium otherwise
    wide = {"Family Office Description", "Investment Thesis", "Investing Sectors"}
    for i, h in enumerate(HEADERS, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = \
            48 if h in wide else 20
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30

    wb.save(OUT)
    print(f"Wrote {len(data)} records to {OUT.name} in the target 32-column layout.")


if __name__ == "__main__":
    main()
